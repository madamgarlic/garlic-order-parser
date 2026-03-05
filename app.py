import streamlit as st
import pandas as pd
import io
from datetime import datetime
import re
from typing import Dict, List, Tuple
from dataclasses import dataclass, field
from collections import defaultdict

# 파싱 결과 데이터 클래스
@dataclass
class ParsedProduct:
    original_text: str
    product_name: str
    weight: str
    unit: str
    is_bulk: bool
    category: str

# 패킹 아이템 데이터 클래스
@dataclass
class PackingItem:
    product_name: str
    category: str
    unit_weight: float
    unit: str
    quantity: int
    total_weight: float
    is_bulk: bool
    order_files: List[str] = field(default_factory=list)

@dataclass
class PackingSummary:
    total_items: int
    total_weight: float
    bulk_items: int
    bulk_weight: float
    regular_items: int
    regular_weight: float
    categories: Dict[str, int]
    total_orders: int
    combined_delivery_count: int
    unique_delivery_locations: int

@dataclass
class DeliveryInfo:
    recipient_name: str
    address: str
    order_count: int
    order_files: List[str] = field(default_factory=list)

class GarlicOrderParser:
    """마늘 주문서 파싱 엔진"""
    
    def __init__(self):
        self.parsing_rules = {
            'primary_target_columns': ['옵션정보', '옵션'],
            'fallback_columns': ['상품명', '제품명', '품목명'],
            'bracket_exceptions': ['특', '대', '중', '소', '대 꼭지제거'],
            'bulk_threshold': 10.0,
        }
    
    def find_target_columns(self, df: pd.DataFrame) -> Tuple[str, str]:
        """타겟 컬럼들 찾기"""
        primary_col = None
        fallback_col = None
        
        # 주요 타겟 컬럼 찾기
        for col in df.columns:
            col_str = str(col).strip()
            for target in self.parsing_rules['primary_target_columns']:
                if target == col_str or target in col_str:
                    primary_col = col
                    break
            if primary_col:
                break
        
        # 보조 타겟 컬럼 찾기
        for col in df.columns:
            col_str = str(col).strip()
            for target in self.parsing_rules['fallback_columns']:
                if target == col_str or target in col_str:
                    fallback_col = col
                    break
            if fallback_col:
                break
        
        if not primary_col and not fallback_col:
            raise ValueError("파싱할 수 있는 컬럼을 찾을 수 없습니다.")
        
        return primary_col, fallback_col
    
    def get_parsing_text(self, row: pd.Series, primary_col: str, fallback_col: str) -> Tuple[str, str]:
        """파싱할 텍스트 가져오기"""
        source_column = ""
        text_to_parse = ""
        
        # 주요 컬럼에서 먼저 확인
        if primary_col and primary_col in row.index:
            primary_text = row[primary_col]
            if pd.notna(primary_text) and str(primary_text).strip():
                text_to_parse = str(primary_text).strip()
                source_column = primary_col
                return text_to_parse, source_column
        
        # 보조 컬럼에서 가져오기
        if fallback_col and fallback_col in row.index:
            fallback_text = row[fallback_col]
            if pd.notna(fallback_text) and str(fallback_text).strip():
                text_to_parse = str(fallback_text).strip()
                source_column = fallback_col
                return text_to_parse, source_column
        
        return "", ""
    
    def clean_text(self, text: str) -> str:
        """기본 텍스트 정제"""
        if pd.isna(text):
            return ""
        
        text = str(text).strip()
        
        # 무조건 삭제할 단어들 제거
        words_to_remove = ['[마늘귀신]', '경북', '의성', '국내산', '마늘귀신', '국산']
        for word in words_to_remove:
            text = re.sub(re.escape(word), '', text, flags=re.IGNORECASE)
        
        # 기존 업소용 표기 삭제
        text = re.sub(r'업소용', '', text, flags=re.IGNORECASE)
        text = re.sub(r'\(업소용\)', '', text, flags=re.IGNORECASE)
        text = re.sub(r'\(\s*업소용\s*\)', '', text, flags=re.IGNORECASE)
        
        # 괄호 처리 (수식 체크 전에 먼저 처리)
        text = self._process_brackets(text)
        
        # 하이픈(-) 처리 추가 - 하이픈 포함 뒤의 모든 정보 삭제
        text = self._process_hyphen(text)
        
        # 손질된 마늘쫑 관련 표현들을 "마늘쫑"으로 대체 및 중복 제거
        garlic_stem_patterns = [
            r'손질된\s*마늘쫑',
            r'\(\s*손질된\s*마늘쫑\s*\)',
            r'손질마늘쫑',
            r'\(\s*손질마늘쫑\s*\)'
        ]
        
        for pattern in garlic_stem_patterns:
            text = re.sub(pattern, '마늘쫑', text, flags=re.IGNORECASE)
        
        # 마늘쫑 중복 제거
        if '마늘쫑' in text.lower():
            weight_matches = re.findall(r'\d+(?:\.\d+)?\s*(?:KG|kg|키로|G|g|그램)', text, re.IGNORECASE)
            weight_part = weight_matches[0] if weight_matches else ""
            
            if weight_part:
                text = f"마늘쫑 {weight_part}"
            else:
                text = "마늘쫑"
        
        # 콜론/슬래시 처리
        text = self._process_delimiters(text)
        
        # 수식 합산 처리 (단위가 같은 경우만)
        text = self._process_math_expressions(text)
        
        # 불필요한 특수문자 제거 (수식 처리 후)
        text = re.sub(r'[,]+', ' ', text)
        
        # 중복 공백 제거
        text = re.sub(r'\s+', ' ', text).strip()
        
        return text
    
    def _process_brackets(self, text: str) -> str:
        """괄호 처리 로직"""
        bracket_pattern = r'\(([^)]+)\)'
        
        def replace_bracket(match):
            content = match.group(1).strip()
            if content in self.parsing_rules['bracket_exceptions']:
                return f" {content} "
            return ' '
        
        result = re.sub(bracket_pattern, replace_bracket, text)
        return re.sub(r'\s+', ' ', result).strip()
    
    def _process_hyphen(self, text: str) -> str:
        """하이픈(-) 처리 로직"""
        if '-' in text:
            text = text.split('-')[0].strip()
        return text
    
    def _process_math_expressions(self, text: str) -> str:
        """수식 합산 처리 (단위가 같은 경우만) - 완전 복원"""
        
        # 무게 단위 수식 (KG, kg, 키로, 키, k)
        kg_patterns = [
            r'(\d+(?:\.\d+)?)\s*(?:KG|kg|키로|키|k)\s*[+x×]\s*(\d+(?:\.\d+)?)\s*(?:KG|kg|키로|키|k)',
        ]
        
        for pattern in kg_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                num1, num2 = float(match.group(1)), float(match.group(2))
                result = num1 + num2
                if result == int(result):
                    result_str = f"{int(result)}KG"
                else:
                    result_str = f"{result}KG"
                text = re.sub(pattern, result_str, text, flags=re.IGNORECASE)
                break
        
        # 그램 단위 수식 (G, g, 그램)
        g_patterns = [
            r'(\d+(?:\.\d+)?)\s*(?:G|g|그램)\s*[+x×]\s*(\d+(?:\.\d+)?)\s*(?:G|g|그램)',
        ]
        
        for pattern in g_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                num1, num2 = float(match.group(1)), float(match.group(2))
                result = num1 + num2
                if result == int(result):
                    result_str = f"{int(result)}G"
                else:
                    result_str = f"{result}G"
                text = re.sub(pattern, result_str, text, flags=re.IGNORECASE)
                break
        
        # 팩 단위 수식
        pack_pattern = r'(\d+)\s*팩\s*[+x×]\s*(\d+)\s*팩'
        pack_match = re.search(pack_pattern, text)
        if pack_match:
            num1, num2 = int(pack_match.group(1)), int(pack_match.group(2))
            result = num1 + num2
            text = re.sub(pack_pattern, f'{result}팩', text)
        
        # 개 단위 수식
        ea_pattern = r'(\d+)\s*개\s*[+x×]\s*(\d+)\s*개'
        ea_match = re.search(ea_pattern, text)
        if ea_match:
            num1, num2 = int(ea_match.group(1)), int(ea_match.group(2))
            result = num1 + num2
            text = re.sub(ea_pattern, f'{result}개', text)
        
        # 포 단위 수식
        po_pattern = r'(\d+)\s*포\s*[+x×]\s*(\d+)\s*포'
        po_match = re.search(po_pattern, text)
        if po_match:
            num1, num2 = int(po_match.group(1)), int(po_match.group(2))
            result = num1 + num2
            text = re.sub(po_pattern, f'{result}포', text)
        
        # 봉 단위 수식
        bag_pattern = r'(\d+)\s*봉\s*[+x×]\s*(\d+)\s*봉'
        bag_match = re.search(bag_pattern, text)
        if bag_match:
            num1, num2 = int(bag_match.group(1)), int(bag_match.group(2))
            result = num1 + num2
            text = re.sub(bag_pattern, f'{result}봉', text)
        
        # 통 단위 수식
        container_pattern = r'(\d+)\s*통\s*[+x×]\s*(\d+)\s*통'
        container_match = re.search(container_pattern, text)
        if container_match:
            num1, num2 = int(container_match.group(1)), int(container_match.group(2))
            result = num1 + num2
            text = re.sub(container_pattern, f'{result}통', text)
        
        return text
    
    def _process_delimiters(self, text: str) -> str:
        """콜론/슬래시 처리"""
        if ':' in text:
            text = text.split(':')[-1].strip()
        
        if '/' in text:
            text = text.split('/')[0].strip()
        
        return text
    
    def extract_weight_info(self, text: str) -> Tuple[str, str, str]:
        """무게 정보 추출 - 완전 복원"""
        weight_patterns = [
            r'(\d+(?:\.\d+)?)\s*키로\s*(?:그램)?',
            r'(\d+(?:\.\d+)?)\s*(KG|kg)',
            r'(\d+(?:\.\d+)?)\s*(G|g|그램)',
            r'(\d+(?:\.\d+)?)\s*(?:키|k)',
        ]
        
        weight_value = ""
        unit = ""
        
        for i, pattern in enumerate(weight_patterns):
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                weight_num = float(match.group(1))
                
                if i == 0:  # 키로
                    weight_value = str(weight_num)
                    unit = 'KG'
                elif i == 1:  # KG/kg
                    weight_value = str(weight_num)
                    unit = 'KG'
                elif i == 2:  # G/g/그램
                    if weight_num >= 1000:
                        weight_value = str(weight_num / 1000)
                        unit = 'KG'
                    else:
                        weight_value = str(weight_num / 1000)
                        unit = 'KG'
                elif i == 3:  # 키/k
                    weight_value = str(weight_num)
                    unit = 'KG'
                
                break
        
        return text, weight_value, unit
    
    def classify_product(self, text: str) -> str:
        """상품 카테고리 분류"""
        text_lower = text.lower()
        
        category_priority = [
            ('마늘쫑', ['마늘쫑', '쫑', '마늘종', '마늘줄기']),
            ('다진마늘', ['다진마늘', '마늘다진', '다진', '으깬마늘', '갈은마늘']),
            ('깐마늘', ['깔마늘', '깐마늘', '마늘깐것', '벗긴마늘', '껍질벗긴마늘']),
            ('닭발', ['닭발'])
        ]
        
        for category, keywords in category_priority:
            if any(keyword in text_lower for keyword in keywords):
                return category
        
        if '마늘' in text_lower:
            return "마늘기타"
        
        return "기타"
    
    def apply_business_rules(self, parsed: ParsedProduct) -> ParsedProduct:
        """비즈니스 규칙 적용"""
        # 깐마늘/다진마늘에서 육쪽이 없으면 대서 자동 추가
        if parsed.category in ['깐마늘', '다진마늘']:
            product_lower = parsed.product_name.lower()
            if '육쪽' not in product_lower:
                if '대서' not in product_lower:
                    parsed.product_name = f"대서 {parsed.product_name}"
        
        # 다진마늘에서 꼭지포함 처리
        if parsed.category == '다진마늘':
            product_lower = parsed.product_name.lower()
            if '꼭지포함' in product_lower:
                parsed.product_name = re.sub(
                    r'꼭지포함', 
                    '* 꼭 지 포 함 *', 
                    parsed.product_name, 
                    flags=re.IGNORECASE
                )
        
        # 업소용 태그 처리 (5KG 이상, 깐마늘/다진마늘만 대상)
        if parsed.category in ['깐마늘', '다진마늘'] and parsed.weight:
            try:
                weight_val = float(parsed.weight)
                if weight_val >= self.parsing_rules['bulk_threshold']:
                    if '업소용' not in parsed.product_name and '** 업 소 용 **' not in parsed.product_name:
                        parsed.product_name = f"** 업 소 용 ** {parsed.product_name}"
                        parsed.is_bulk = True
            except (ValueError, TypeError):
                pass
        
        # 상품명 정리
        parsed.product_name = re.sub(r'\s+', ' ', parsed.product_name).strip()
        
        return parsed
    
    def parse_single_item(self, text: str) -> Tuple[ParsedProduct, List[str]]:
        """단일 항목 파싱"""
        original_text = text
        warnings = []
        
        try:
            # 기본 정제
            cleaned_text = self.clean_text(text)
            
            # 무게 정보 추출 (원본 텍스트 보존)
            _, weight, unit = self.extract_weight_info(cleaned_text)
            
            # 카테고리 분류
            category = self.classify_product(cleaned_text)
            
            # ParsedProduct 객체 생성
            parsed = ParsedProduct(
                original_text=original_text,
                product_name=cleaned_text,
                weight=weight,
                unit=unit,
                is_bulk=False,
                category=category
            )
            
            # 비즈니스 규칙 적용
            parsed = self.apply_business_rules(parsed)
            
            return parsed, warnings
            
        except Exception as e:
            warnings.append(f"파싱 실패: {str(e)}")
            
            return ParsedProduct(
                original_text=original_text,
                product_name=text,
                weight="",
                unit="",
                is_bulk=False,
                category="오류"
            ), warnings
    
    def parse_dataframe(self, df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict]:
        """데이터프레임 전체 파싱"""
        result_df = df.copy()
        parsing_log = {
            'total_rows': len(df),
            'success_count': 0,
            'warning_count': 0,
            'error_count': 0,
            'warnings': [],
            'errors': [],
            'primary_column': None,
            'fallback_column': None,
            'source_stats': {'primary': 0, 'fallback': 0, 'empty': 0}
        }
        
        try:
            # 타겟 컬럼들 찾기
            primary_col, fallback_col = self.find_target_columns(df)
            parsing_log['primary_column'] = primary_col
            parsing_log['fallback_column'] = fallback_col
            
            # 결과 컬럼명 결정
            result_col_base = primary_col if primary_col else fallback_col
            
            # 파싱 결과 저장할 새 컬럼들
            result_df[f'{result_col_base}_parsed'] = ""
            result_df[f'{result_col_base}_weight'] = ""
            result_df[f'{result_col_base}_unit'] = ""
            result_df[f'{result_col_base}_category'] = ""
            result_df[f'{result_col_base}_is_bulk'] = False
            result_df[f'{result_col_base}_source'] = ""
            
            # 각 행 파싱
            for idx, row in df.iterrows():
                # 파싱할 텍스트 가져오기
                text_to_parse, source_column = self.get_parsing_text(row, primary_col, fallback_col)
                
                if not text_to_parse:
                    parsing_log['source_stats']['empty'] += 1
                    continue
                
                # 소스 통계 업데이트
                if source_column == primary_col:
                    parsing_log['source_stats']['primary'] += 1
                elif source_column == fallback_col:
                    parsing_log['source_stats']['fallback'] += 1
                
                # 파싱 실행
                parsed, warnings = self.parse_single_item(text_to_parse)
                
                # 결과 저장
                result_df.loc[idx, f'{result_col_base}_parsed'] = parsed.product_name
                result_df.loc[idx, f'{result_col_base}_weight'] = parsed.weight
                result_df.loc[idx, f'{result_col_base}_unit'] = parsed.unit
                result_df.loc[idx, f'{result_col_base}_category'] = parsed.category
                result_df.loc[idx, f'{result_col_base}_is_bulk'] = parsed.is_bulk
                result_df.loc[idx, f'{result_col_base}_source'] = source_column
                
                # 로그 집계
                if warnings:
                    parsing_log['warning_count'] += 1
                    parsing_log['warnings'].extend([f"행 {idx+1} ({source_column}): {w}" for w in warnings])
                else:
                    parsing_log['success_count'] += 1
            
            # 원본 컬럼을 파싱된 결과로 대체
            if primary_col:
                result_df[primary_col] = result_df[f'{result_col_base}_parsed']
            elif fallback_col:
                result_df[fallback_col] = result_df[f'{result_col_base}_parsed']
            
        except Exception as e:
            parsing_log['error_count'] += 1
            parsing_log['errors'].append(f"전체 파싱 실패: {str(e)}")
        
        return result_df, parsing_log

class PackingListGenerator:
    """패킹리스트 생성기"""
    
    def __init__(self):
        self.aggregation_data = defaultdict(lambda: {
            'quantity': 0,
            'total_weight': 0.0,
            'order_files': set(),
            'unit_weight': 0.0,
            'unit': 'KG',
            'category': '',
            'is_bulk': False
        })
        self.delivery_data = defaultdict(lambda: {
            'order_count': 0,
            'order_files': set()
        })
        self.total_orders_processed = 0
    
    def is_bulk_product(self, product_name: str, is_bulk_flag: bool) -> bool:
        """업소용 상품 판단"""
        if is_bulk_flag:
            return True
        return '업소용' in product_name or '** 업 소 용 **' in product_name
    
    def generate_aggregation_key(self, product_name: str, weight: str, category: str, is_bulk: bool) -> str:
        """집계 키 생성"""
        
        # 업소용 상품이나 마늘쫑의 경우: 상품명 그대로
        if self.is_bulk_product(product_name, is_bulk) or category == '마늘쫑':
            return f"{category}_{product_name}"
        
        # 닭발의 경우: 수량만 집계 (무게 정보 무시)
        elif category == '닭발':
            clean_name = self._remove_weight_from_name(product_name)
            return f"{category}_{clean_name}"
        
        # 일반 상품: 무게 제거한 상품명으로 그룹화
        else:
            clean_name = self._remove_weight_from_name(product_name)
            return f"{category}_{clean_name}"
    
    def _remove_weight_from_name(self, product_name: str) -> str:
        """상품명에서 무게 정보 제거"""
        weight_patterns = [
            r'\d+(?:\.\d+)?\s*키로\s*(?:그램)?',
            r'\d+(?:\.\d+)?\s*(?:KG|kg)',
            r'\d+(?:\.\d+)?\s*(?:G|g|그램)',
            r'\d+(?:\.\d+)?\s*(?:키|k)',
            r'\d+(?:\.\d+)?\s*(?:킬로|kilo)',
        ]
        
        clean_name = product_name
        for pattern in weight_patterns:
            clean_name = re.sub(pattern, '', clean_name, flags=re.IGNORECASE)
        
        # 불필요한 공백과 기호 정리
        clean_name = re.sub(r'\s+', ' ', clean_name).strip()
        # 앞뒤 기호 제거
        clean_name = re.sub(r'^[,.\-\s]+|[,.\-\s]+$', '', clean_name).strip()
        
        return clean_name
    
    def find_delivery_columns(self, df: pd.DataFrame) -> Tuple[str, str]:
        """배송 관련 컬럼 찾기"""
        recipient_col = None
        address_col = None
        
        # 수령인명 컬럼 찾기
        for col in df.columns:
            col_str = str(col).strip()
            if '수령인명' in col_str or '받는분' in col_str or '수령인' in col_str:
                recipient_col = col
                break
        
        # 주소 컬럼 찾기
        for col in df.columns:
            col_str = str(col).strip()
            if '주소' in col_str or '배송지' in col_str or '배송주소' in col_str:
                address_col = col
                break
        
        return recipient_col, address_col
    
    def add_order_data(self, df: pd.DataFrame, source_file: str):
        """주문 데이터 추가"""
        # 파싱된 컬럼 찾기
        parsed_cols = {}
        for col in df.columns:
            if col.endswith('_parsed'):
                base_name = col[:-7]
                parsed_cols['parsed'] = col
                parsed_cols['weight'] = f"{base_name}_weight"
                parsed_cols['category'] = f"{base_name}_category"
                parsed_cols['is_bulk'] = f"{base_name}_is_bulk"
                break
        
        if not parsed_cols:
            st.error("파싱된 컬럼을 찾을 수 없습니다.")
            return
        
        # 수량 컬럼 찾기
        quantity_col = None
        for col in df.columns:
            if any(keyword in col.lower() for keyword in ['수량', 'qty', 'quantity', '개수']):
                quantity_col = col
                break
        
        if not quantity_col:
            st.error("수량 컬럼을 찾을 수 없습니다.")
            return
        
        # 배송 관련 컬럼 찾기
        recipient_col, address_col = self.find_delivery_columns(df)
        
        # 데이터 집계
        for idx, row in df.iterrows():
            try:
                product_name = str(row[parsed_cols['parsed']])
                weight = str(row[parsed_cols['weight']])
                category = str(row[parsed_cols['category']])
                is_bulk = bool(row[parsed_cols['is_bulk']])
                quantity = int(row[quantity_col]) if pd.notna(row[quantity_col]) else 0
                
                if quantity <= 0 or not product_name or product_name == 'nan':
                    continue
                
                # 전체 주문 건수 증가
                self.total_orders_processed += 1
                
                # 집계 키 생성
                agg_key = self.generate_aggregation_key(product_name, weight, category, is_bulk)
                
                # 상품 데이터 집계
                agg_data = self.aggregation_data[agg_key]
                agg_data['category'] = category
                agg_data['is_bulk'] = self.is_bulk_product(product_name, is_bulk)
                agg_data['order_files'].add(source_file)
                
                # 무게와 수량 처리
                if weight and weight != 'nan':
                    unit_weight = float(weight)
                    agg_data['unit_weight'] = unit_weight
                    
                    # 업소용, 마늘쫑: 수량만 합산
                    if self.is_bulk_product(product_name, is_bulk) or category == '마늘쫑':
                        agg_data['quantity'] += quantity
                        agg_data['total_weight'] += unit_weight * quantity
                    # 닭발: 수량만 합산, 무게 무시
                    elif category == '닭발':
                        agg_data['quantity'] += quantity
                        agg_data['total_weight'] = 0
                    # 일반 상품: 무게 X 수량을 1KG로 나눈 값
                    else:
                        final_quantity = (unit_weight * quantity) / 1.0
                        agg_data['quantity'] += final_quantity
                        agg_data['total_weight'] += unit_weight * quantity
                else:
                    # 무게 정보가 없는 경우 (닭발 등)
                    agg_data['quantity'] += quantity
                    if category == '닭발':
                        agg_data['total_weight'] = 0
                
                # 배송 정보 집계
                if recipient_col and address_col and recipient_col in df.columns and address_col in df.columns:
                    recipient = str(row[recipient_col]) if pd.notna(row[recipient_col]) else ""
                    address = str(row[address_col]) if pd.notna(row[address_col]) else ""
                    
                    if recipient and address and recipient != 'nan' and address != 'nan':
                        delivery_key = f"{recipient}_{address}"
                        delivery_data = self.delivery_data[delivery_key]
                        delivery_data['order_count'] += 1
                        delivery_data['order_files'].add(source_file)
                
            except Exception as e:
                continue
    
    def generate_packing_list(self) -> Tuple[List[PackingItem], PackingSummary]:
        """패킹리스트 생성"""
        packing_items = []
        
        for agg_key, agg_data in self.aggregation_data.items():
            if agg_data['quantity'] > 0:
                key_parts = agg_key.split('_', 1)
                category = key_parts[0]
                raw_product_name = key_parts[1] if len(key_parts) > 1 else ""
                
                # 표시할 상품명 결정
                if agg_data['is_bulk'] or category == '마늘쫑':
                    display_name = raw_product_name
                    display_quantity = int(agg_data['quantity'])
                else:
                    display_name = raw_product_name
                    display_quantity = agg_data['quantity']
                
                packing_item = PackingItem(
                    product_name=display_name,
                    category=category,
                    unit_weight=agg_data['unit_weight'],
                    unit=agg_data['unit'],
                    quantity=display_quantity,
                    total_weight=agg_data['total_weight'],
                    is_bulk=agg_data['is_bulk'],
                    order_files=list(agg_data['order_files'])
                )
                
                packing_items.append(packing_item)
        
        # 패킹리스트 정렬 개선: 같은 제품끼리 그룹핑
        packing_items = sorted(packing_items, key=lambda x: (
            x.category,        # 1순위: 카테고리
            x.product_name     # 2순위: 상품명 (같은 제품끼리 묶임)
        ))
        
        # 요약 정보 생성
        summary = self._generate_summary(packing_items)
        
        return packing_items, summary
    
    def _generate_summary(self, items: List[PackingItem]) -> PackingSummary:
        """요약 정보 생성"""
        total_items = len(items)
        total_weight = sum(item.total_weight for item in items)
        
        bulk_items = [item for item in items if item.is_bulk]
        bulk_count = len(bulk_items)
        bulk_weight = sum(item.total_weight for item in bulk_items)
        
        regular_count = total_items - bulk_count
        regular_weight = total_weight - bulk_weight
        
        categories = defaultdict(int)
        for item in items:
            categories[item.category] += item.quantity
        
        # 합배송 분석
        combined_deliveries = sum(1 for delivery_data in self.delivery_data.values() if delivery_data['order_count'] > 1)
        unique_deliveries = len(self.delivery_data)
        
        return PackingSummary(
            total_items=total_items,
            total_weight=total_weight,
            bulk_items=bulk_count,
            bulk_weight=bulk_weight,
            regular_items=regular_count,
            regular_weight=regular_weight,
            categories=dict(categories),
            total_orders=self.total_orders_processed,
            combined_delivery_count=combined_deliveries,
            unique_delivery_locations=unique_deliveries
        )
    
    def get_statistics(self) -> Dict:
        """현재 집계 통계 반환"""
        total_orders = sum(data['quantity'] for data in self.aggregation_data.values())
        total_products = len(self.aggregation_data)
        total_weight = sum(data['total_weight'] for data in self.aggregation_data.values())
        
        categories = defaultdict(int)
        for data in self.aggregation_data.values():
            categories[data['category']] += data['quantity']
        
        # 합배송 통계
        combined_deliveries = sum(1 for delivery_data in self.delivery_data.values() if delivery_data['order_count'] > 1)
        unique_deliveries = len(self.delivery_data)
        
        return {
            'total_orders': total_orders,
            'total_products': total_products,
            'total_weight': total_weight,
            'categories': dict(categories),
            'files_processed': len(set().union(*(data['order_files'] for data in self.aggregation_data.values() if data['order_files']))),
            'total_orders_processed': self.total_orders_processed,
            'combined_deliveries': combined_deliveries,
            'unique_deliveries': unique_deliveries
        }
    
    def get_delivery_analysis(self) -> List[DeliveryInfo]:
        """배송 분석 결과 반환"""
        delivery_list = []
        
        for delivery_key, delivery_data in self.delivery_data.items():
            if '_' in delivery_key:
                recipient, address = delivery_key.split('_', 1)
                delivery_info = DeliveryInfo(
                    recipient_name=recipient,
                    address=address,
                    order_count=delivery_data['order_count'],
                    order_files=list(delivery_data['order_files'])
                )
                delivery_list.append(delivery_info)
        
        # 주문 건수 순으로 정렬 (많은 순)
        delivery_list.sort(key=lambda x: x.order_count, reverse=True)
        
        return delivery_list
    
    def clear_data(self):
        """집계 데이터 초기화"""
        self.aggregation_data.clear()
        self.delivery_data.clear()
        self.total_orders_processed = 0

# 색상 표시 관련 함수들 - 완전 복원
def find_combined_delivery_rows(df: pd.DataFrame) -> list:
    """합배송 행들의 인덱스 찾기 (노란색 색칠용)"""
    try:
        # 배송 관련 컬럼 찾기
        recipient_col = None
        address_col = None
        phone_col = None
        
        for col in df.columns:
            col_clean = re.sub(r'[\s\-_]+', '', str(col).strip())
            
            if not recipient_col:
                recipient_keywords = ['수취인명', '수취인', '수령인명']
                if col_clean in recipient_keywords:
                    recipient_col = col
            
            if not address_col:
                address_keywords = ['주소', '기본배송지', '배송지주소']
                if col_clean in address_keywords:
                    address_col = col
            
            if not phone_col:
                phone_keywords = ['수령인연락처1', '수취인연락처1', '수령인휴대폰', '휴대폰번호']
                if col_clean in phone_keywords:
                    phone_col = col
        
        # 3개 컬럼이 모두 있어야 합배송 체크 가능
        if not (recipient_col and address_col and phone_col):
            return []
        
        # 배송 그룹 생성
        delivery_groups = {}
        
        for idx, row in df.iterrows():
            recipient = str(row[recipient_col]) if pd.notna(row[recipient_col]) else ""
            address = str(row[address_col]) if pd.notna(row[address_col]) else ""
            phone = str(row[phone_col]) if pd.notna(row[phone_col]) else ""
            
            # 유효한 데이터만 처리
            if recipient and address and phone and recipient != 'nan' and address != 'nan' and phone != 'nan':
                delivery_key = f"{recipient}_{address}_{phone}"
                
                if delivery_key not in delivery_groups:
                    delivery_groups[delivery_key] = []
                delivery_groups[delivery_key].append(idx)
        
        # 합배송 행들 찾기 (2개 이상의 주문이 있는 그룹)
        combined_rows = []
        for group_rows in delivery_groups.values():
            if len(group_rows) > 1:  # 합배송인 경우
                combined_rows.extend(group_rows)
        
        return combined_rows
        
    except Exception as e:
        return []

def find_heavy_order_rows(df: pd.DataFrame) -> list:
    """10kg 초과 주문 행들의 인덱스 찾기 (연두색 색칠용)"""
    try:
        heavy_rows = []
        
        # 수량 컬럼 찾기
        quantity_col = None
        for col in df.columns:
            if any(keyword in col.lower() for keyword in ['수량', 'qty', 'quantity', '개수']):
                quantity_col = col
                break
        
        # 무게 컬럼 찾기
        weight_col = None
        for col in df.columns:
            if col.endswith('_weight'):
                weight_col = col
                break
        
        if not quantity_col or not weight_col:
            return []
        
        # 각 행에서 10kg 초과 체크
        for idx, row in df.iterrows():
            try:
                quantity = pd.to_numeric(row[quantity_col], errors='coerce')
                unit_weight = pd.to_numeric(row[weight_col], errors='coerce')
                
                if pd.notna(quantity) and pd.notna(unit_weight):
                    total_weight = quantity * unit_weight
                    if total_weight > 10.0:
                        heavy_rows.append(idx)
            except:
                continue
        
        return heavy_rows
        
    except Exception as e:
        return []

# 정렬 함수들 - 완전 복원
def apply_sorting_to_parsed_file(df: pd.DataFrame) -> pd.DataFrame:
    """파싱된 파일에 정렬 적용 - 전체 행 데이터 함께 정렬 (데이터 무결성 보장)"""
    try:
        # 원본 DataFrame의 완전한 복사본 생성 (모든 데이터 보존)
        df_original = df.copy()
        
        # 필요한 컬럼들 찾기
        parsed_col = None
        quantity_col = None
        
        # 파싱된 상품명 컬럼 찾기
        for col in df_original.columns:
            if col.endswith('_parsed'):
                parsed_col = col
                break
        
        # 수량 컬럼 찾기
        for col in df_original.columns:
            if any(keyword in col.lower() for keyword in ['수량', 'qty', 'quantity', '개수']):
                quantity_col = col
                break
        
        if not parsed_col or not quantity_col:
            st.warning(f"⚠️ 정렬에 필요한 컬럼을 찾을 수 없습니다. 파싱컬럼: {parsed_col}, 수량컬럼: {quantity_col}")
            return df_original
        
        # 배송 관련 컬럼 찾기
        recipient_col = None
        address_col = None
        phone_col = None
        
        for col in df_original.columns:
            col_clean = re.sub(r'[\s\-_]+', '', str(col).strip())
            
            if not recipient_col:
                recipient_keywords = ['수취인명', '수취인', '수령인명']
                if col_clean in recipient_keywords:
                    recipient_col = col
            
            if not address_col:
                address_keywords = ['주소', '기본배송지', '배송지주소']
                if col_clean in address_keywords:
                    address_col = col
            
            if not phone_col:
                phone_keywords = ['수령인연락처1', '수취인연락처1', '수령인휴대폰', '휴대폰번호']
                if col_clean in phone_keywords:
                    phone_col = col
        
        # 간단한 정보만 출력
        st.info(f"🔧 정렬 기준: {parsed_col} (수량: {quantity_col})")
        
        # ⭐ 핵심: 원본 인덱스를 보존하면서 정렬용 컬럼만 추가
        df_for_sorting = df_original.copy()
        
        # 합배송 여부 판단
        has_combined_delivery = False
        
        if recipient_col and address_col and phone_col:
            # 배송 키 생성
            df_for_sorting['_temp_delivery_key'] = (
                df_for_sorting[recipient_col].astype(str) + '_' + 
                df_for_sorting[address_col].astype(str) + '_' + 
                df_for_sorting[phone_col].astype(str)
            )
            
            # 합배송 건수 계산
            delivery_counts = df_for_sorting['_temp_delivery_key'].value_counts()
            df_for_sorting['_temp_is_combined'] = df_for_sorting['_temp_delivery_key'].map(lambda x: delivery_counts[x] > 1)
            
            # 실제 합배송이 있는지 확인
            has_combined_delivery = df_for_sorting['_temp_is_combined'].any()
            
            if has_combined_delivery:
                combined_count = delivery_counts[delivery_counts > 1].sum()
                st.success(f"🚚 합배송 감지: {combined_count}건 → 상단 고정 정렬")
        
        # ⭐ 10kg 초과 주문 감지 (연두색 표시용)
        df_for_sorting['_temp_is_heavy'] = False
        if quantity_col in df_for_sorting.columns:
            try:
                for idx, row in df_for_sorting.iterrows():
                    quantity = pd.to_numeric(row[quantity_col], errors='coerce')
                    # 파싱된 무게 정보 찾기
                    weight_col = None
                    for col in df_for_sorting.columns:
                        if col.endswith('_weight'):
                            weight_col = col
                            break
                    
                    if weight_col and pd.notna(row[weight_col]) and pd.notna(quantity):
                        unit_weight = pd.to_numeric(row[weight_col], errors='coerce')
                        total_weight = quantity * unit_weight
                        if total_weight > 10.0:
                            df_for_sorting.loc[idx, '_temp_is_heavy'] = True
            except:
                pass  # 오류 시 무시
        
        heavy_count = df_for_sorting['_temp_is_heavy'].sum()
        if heavy_count > 0:
            st.info(f"🟢 10kg 초과 주문 감지: {heavy_count}건 → 연두색 표시")
        
        # 수량을 숫자로 변환 (정렬용)
        df_for_sorting['_temp_quantity_sort'] = pd.to_numeric(df_for_sorting[quantity_col], errors='coerce').fillna(0)
        
        # ⭐⭐⭐ 수정된 정렬: 제품명 → 수량 순으로 변경
        if has_combined_delivery:
            # 합배송이 있는 경우: 합배송 상단 고정 + 제품명 + 수량순
            sort_columns = ['_temp_is_combined', parsed_col, '_temp_quantity_sort']
            sort_ascending = [False, True, True]  # 합배송 상단, 제품명 가나다순, 수량 적은순
            st.info("📋 정렬 순서: 합배송 상단 고정 → 제품명 가나다순 → 수량 적은순")
        else:
            # 합배송이 없는 경우: 제품명 + 수량순
            sort_columns = [parsed_col, '_temp_quantity_sort']
            sort_ascending = [True, True]  # 제품명 가나다순, 수량 적은순
            st.info("📋 정렬 순서: 제품명 가나다순 → 수량 적은순")
        
        # ✅ 전체 정렬 실행 - 모든 원본 데이터가 행별로 함께 움직임 (데이터 무결성 보장)
        sorted_df = df_for_sorting.sort_values(
            sort_columns,
            ascending=sort_ascending,
            na_position='last'
        ).reset_index(drop=True)
        
        # 임시 정렬용 컬럼들만 제거 (원본 데이터는 모두 보존)
        temp_columns = ['_temp_delivery_key', '_temp_is_combined', '_temp_quantity_sort', '_temp_is_heavy']
        for temp_col in temp_columns:
            if temp_col in sorted_df.columns:
                sorted_df = sorted_df.drop(columns=[temp_col])
        
        # 데이터 무결성 검증
        original_shape = df_original.shape
        sorted_shape = sorted_df.shape
        
        if original_shape != sorted_shape:
            st.error(f"❌ 데이터 무결성 오류: 원본 {original_shape} ≠ 정렬후 {sorted_shape}")
            return df_original
        
        st.success(f"✅ 전체 {len(sorted_df)}행 × {len(sorted_df.columns)}열 정렬 완료 (모든 데이터 동기화)")
        
        return sorted_df
        
    except Exception as e:
        st.error(f"정렬 처리 중 오류: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return df

def apply_sorting_to_parsed_file_silent(df: pd.DataFrame) -> pd.DataFrame:
    """정렬 함수의 조용한 버전 (일괄 처리용 - 출력 없음)"""
    try:
        # 원본 DataFrame의 완전한 복사본 생성
        df_original = df.copy()
        
        # 필요한 컬럼들 찾기
        parsed_col = None
        quantity_col = None
        
        for col in df_original.columns:
            if col.endswith('_parsed'):
                parsed_col = col
                break
        
        for col in df_original.columns:
            if any(keyword in col.lower() for keyword in ['수량', 'qty', 'quantity', '개수']):
                quantity_col = col
                break
        
        if not parsed_col or not quantity_col:
            return df_original
        
        # 배송 관련 컬럼 찾기
        recipient_col = None
        address_col = None
        phone_col = None
        
        for col in df_original.columns:
            col_clean = re.sub(r'[\s\-_]+', '', str(col).strip())
            
            if not recipient_col:
                recipient_keywords = ['수취인명', '수취인', '수령인명']
                if col_clean in recipient_keywords:
                    recipient_col = col
            
            if not address_col:
                address_keywords = ['주소', '기본배송지', '배송지주소']
                if col_clean in address_keywords:
                    address_col = col
            
            if not phone_col:
                phone_keywords = ['수령인연락처1', '수취인연락처1', '수령인휴대폰', '휴대폰번호']
                if col_clean in phone_keywords:
                    phone_col = col
        
        # 정렬용 DataFrame 생성 (모든 원본 데이터 포함)
        df_for_sorting = df_original.copy()
        
        # 합배송 처리
        has_combined_delivery = False
        
        if recipient_col and address_col and phone_col:
            df_for_sorting['_temp_delivery_key'] = (
                df_for_sorting[recipient_col].astype(str) + '_' + 
                df_for_sorting[address_col].astype(str) + '_' + 
                df_for_sorting[phone_col].astype(str)
            )
            
            delivery_counts = df_for_sorting['_temp_delivery_key'].value_counts()
            df_for_sorting['_temp_is_combined'] = df_for_sorting['_temp_delivery_key'].map(lambda x: delivery_counts[x] > 1)
            has_combined_delivery = df_for_sorting['_temp_is_combined'].any()
        
        # ⭐ 10kg 초과 주문 처리 (연두색용)
        df_for_sorting['_temp_is_heavy'] = False
        if quantity_col in df_for_sorting.columns:
            try:
                for idx, row in df_for_sorting.iterrows():
                    quantity = pd.to_numeric(row[quantity_col], errors='coerce')
                    weight_col = None
                    for col in df_for_sorting.columns:
                        if col.endswith('_weight'):
                            weight_col = col
                            break
                    
                    if weight_col and pd.notna(row[weight_col]) and pd.notna(quantity):
                        unit_weight = pd.to_numeric(row[weight_col], errors='coerce')
                        total_weight = quantity * unit_weight
                        if total_weight > 10.0:
                            df_for_sorting.loc[idx, '_temp_is_heavy'] = True
            except:
                pass

        # 수량 변환
        df_for_sorting['_temp_quantity_sort'] = pd.to_numeric(df_for_sorting[quantity_col], errors='coerce').fillna(0)
        
        # ⭐ 수정된 정렬: 제품명 → 수량 순
        if has_combined_delivery:
            sort_columns = ['_temp_is_combined', parsed_col, '_temp_quantity_sort']
            sort_ascending = [False, True, True]
        else:
            sort_columns = [parsed_col, '_temp_quantity_sort']
            sort_ascending = [True, True]
        
        # ✅ 전체 행 정렬 (데이터 무결성 보장)
        sorted_df = df_for_sorting.sort_values(
            sort_columns,
            ascending=sort_ascending,
            na_position='last'
        ).reset_index(drop=True)
        
        # 임시 컬럼들만 제거
        temp_columns = ['_temp_delivery_key', '_temp_is_combined', '_temp_quantity_sort', '_temp_is_heavy']
        for temp_col in temp_columns:
            if temp_col in sorted_df.columns:
                sorted_df = sorted_df.drop(columns=[temp_col])
        
        return sorted_df
        
    except Exception as e:
        return df

# Streamlit 앱
def setup_page():
    """페이지 설정"""
    st.set_page_config(
        page_title="마늘 주문서 자동 파서 & 패킹리스트 생성기",
        page_icon="🧄",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    st.title("🧄 마늘 주문서 자동 파서 및 패킹리스트 생성기 (v3.0)")
    st.markdown("---")

def initialize_session_state():
    """세션 상태 초기화"""
    if 'parser' not in st.session_state:
        st.session_state.parser = GarlicOrderParser()
    
    if 'packing_generator' not in st.session_state:
        st.session_state.packing_generator = PackingListGenerator()
    
    if 'processed_files' not in st.session_state:
        st.session_state.processed_files = {}
    
    if 'parsing_logs' not in st.session_state:
        st.session_state.parsing_logs = {}

def display_sidebar():
    """사이드바 표시"""
    with st.sidebar:
        st.header("📋 처리 현황")
        
        stats = st.session_state.packing_generator.get_statistics()
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("처리된 파일", stats['files_processed'])
            st.metric("전체 주문 건수", stats['total_orders_processed'])
            st.metric("총 상품 종류", stats['total_products'])
        
        with col2:
            st.metric("총 주문 수량", stats['total_orders'])
            st.metric("총 무게(KG)", f"{stats['total_weight']:.1f}")
            st.metric("고유 배송지", stats['unique_deliveries'])
        
        # 합배송 정보
        if stats['combined_deliveries'] > 0:
            st.markdown("### 🚚 배송 정보")
            st.metric("합배송 건", stats['combined_deliveries'])
            st.write(f"• 총 배송지: {stats['unique_deliveries']}곳")
            st.write(f"• 합배송률: {(stats['combined_deliveries']/stats['unique_deliveries']*100):.1f}%" if stats['unique_deliveries'] > 0 else "• 합배송률: 0%")
        
        if stats['categories']:
            st.subheader("📊 카테고리별 현황")
            for category, qty in stats['categories'].items():
                if isinstance(qty, float):
                    st.write(f"• {category}: {qty:.0f}개")
                else:
                    st.write(f"• {category}: {qty}개")
        
        st.markdown("---")
        
        if st.button("🔄 데이터 초기화", type="secondary"):
            st.session_state.packing_generator.clear_data()
            st.session_state.processed_files.clear()
            st.session_state.parsing_logs.clear()
            st.rerun()

def main():
    """메인 애플리케이션"""
    setup_page()
    initialize_session_state()
    
    # 사이드바
    display_sidebar()
    
    # 파일 업로드 섹션
    st.header("📁 1. 엑셀 파일 업로드")
    
    uploaded_files = st.file_uploader(
        "마늘 주문서 엑셀 파일들을 선택하세요",
        type=['xlsx', 'xls'],
        accept_multiple_files=True,
        help="여러 개의 엑셀 파일을 동시에 업로드할 수 있습니다."
    )
    
    if uploaded_files:
        st.success(f"📎 {len(uploaded_files)}개 파일이 업로드되었습니다.")
        
        # 파싱 섹션
        st.header("⚙️ 2. 주문서 파싱")
        
        if st.button("🚀 파싱 시작", type="primary"):
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            for i, file in enumerate(uploaded_files):
                status_text.text(f"처리 중: {file.name}")
                
                try:
                    df = pd.read_excel(file)
                    
                    st.subheader(f"🔍 {file.name} 파싱 결과")
                    
                    parsed_df, parsing_log = st.session_state.parser.parse_dataframe(df)
                    
                    # 파싱 로그 표시
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("성공", parsing_log['success_count'])
                    with col2:
                        st.metric("경고", parsing_log['warning_count'])
                    with col3:
                        st.metric("오류", parsing_log['error_count'])
                    
                    if parsing_log.get('primary_column'):
                        st.success(f"주요 컬럼에서 파싱: {parsing_log['primary_column']}")
                    if parsing_log.get('fallback_column'):
                        st.info(f"보조 컬럼 사용 가능: {parsing_log['fallback_column']}")
                    
                    if parsing_log.get('errors'):
                        st.error("오류 발생:")
                        for error in parsing_log['errors']:
                            st.write(f"  • {error}")
                    
                    st.session_state.processed_files[file.name] = {
                        'original_df': df,
                        'parsed_df': parsed_df,
                        'file_obj': file
                    }
                    st.session_state.parsing_logs[file.name] = parsing_log
                    
                    st.session_state.packing_generator.add_order_data(parsed_df, file.name)
                    
                except Exception as e:
                    st.error(f"❌ {file.name} 처리 실패: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())
                    continue
                
                progress_bar.progress((i + 1) / len(uploaded_files))
            
            status_text.text("✅ 모든 파일 처리 완료!")
            st.success("파싱이 완료되었습니다!")
            st.rerun()
    
    # 파싱 결과 표시
    if st.session_state.processed_files:
        st.header("📊 3. 파싱 결과")
        
        for filename, file_data in st.session_state.processed_files.items():
            with st.expander(f"📄 {filename} 결과"):
                log = st.session_state.parsing_logs.get(filename, {})
                
                col1, col2 = st.columns([3, 1])
                
                with col1:
                    # 파싱 결과 미리보기
                    if st.checkbox(f"미리보기 - {filename}", key=f"preview_{filename}"):
                        original_df = file_data['original_df']
                        parsed_df = file_data['parsed_df']
                        
                        st.markdown("##### 파싱 전/후 비교")
                        compare_cols = st.columns(2)
                        
                        with compare_cols[0]:
                            st.markdown("**파싱 전 (처음 5행)**")
                            st.dataframe(original_df.head(5), use_container_width=True)
                        
                        with compare_cols[1]:
                            st.markdown("**파싱 후 (처음 5행)**")
                            # 주요 컬럼만 표시
                            display_cols = []
                            for col in parsed_df.columns:
                                if any(suffix in col for suffix in ['_parsed', '_weight', '_category', '_is_bulk']) or col in ['수량', '수령인명', '주소']:
                                    display_cols.append(col)
                            
                            if display_cols:
                                st.dataframe(parsed_df[display_cols].head(5), use_container_width=True)
                            else:
                                st.dataframe(parsed_df.head(5), use_container_width=True)
                
                with col2:
                    # 정렬된 파싱 파일 다운로드 (색상 표시 포함) - 완전 복원
                    parsed_df = file_data['parsed_df'].copy()
                    
                    # 정렬 처리
                    sorted_df = apply_sorting_to_parsed_file(parsed_df)
                    
                    # 색상 표시를 위한 엑셀 파일 생성
                    output = io.BytesIO()
                    
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        sorted_df.to_excel(writer, sheet_name='파싱결과_정렬됨', index=False)
                        
                        # 색상 표시 적용
                        workbook = writer.book
                        worksheet = writer.sheets['파싱결과_정렬됨']
                        
                        # 합배송 그룹 찾기 (색칠용)
                        combined_rows = find_combined_delivery_rows(sorted_df)
                        heavy_rows = find_heavy_order_rows(sorted_df)  # 10kg 초과 주문 찾기
                        
                        if combined_rows or heavy_rows:
                            from openpyxl.styles import PatternFill
                            
                            # 색상 정의
                            yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # 합배송 - 노란색
                            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")    # 10kg 초과 - 연두색
                            red_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")      # 합배송+10kg - 연한 빨간색
                            
                            # 겹치는 행들 찾기 (합배송 + 10kg 초과)
                            overlap_rows = list(set(combined_rows) & set(heavy_rows))
                            
                            # 색칠 적용
                            for row_idx in range(len(sorted_df)):
                                excel_row = row_idx + 2  # 헤더 때문에 +2
                                
                                if row_idx in overlap_rows:
                                    # 합배송 + 10kg 초과 = 연한 빨간색
                                    fill = red_fill
                                elif row_idx in combined_rows:
                                    # 합배송만 = 노란색
                                    fill = yellow_fill
                                elif row_idx in heavy_rows:
                                    # 10kg 초과만 = 연두색
                                    fill = green_fill
                                else:
                                    continue
                                
                                # 해당 행의 모든 셀에 색칠
                                for col_idx in range(1, len(sorted_df.columns) + 1):
                                    cell = worksheet.cell(row=excel_row, column=col_idx)
                                    cell.fill = fill
                    
                    output.seek(0)
                    
                    download_filename = f"{filename.split('.')[0]}_parsed_sorted.xlsx"
                    
                    st.download_button(
                        label="📎 정렬된 파싱 결과 다운로드",
                        data=output.getvalue(),
                        file_name=download_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_{filename}"
                    )
                    
                    st.caption("✨ 제품명 → 수량순 자동 정렬")
                    st.caption("🎨 합배송: 노란색 | 10kg초과: 연두색 | 둘다: 연한빨강")
        
        # 전체 파일 일괄 다운로드 섹션 - 완전 복원
        if len(st.session_state.processed_files) > 1:
            st.markdown("---")
            st.subheader("📦 전체 파일 일괄 다운로드")
            
            if st.button("📎 모든 정렬된 파일 한번에 다운로드", type="secondary"):
                try:
                    # 모든 파일을 처리해서 ZIP 생성
                    import zipfile
                    
                    # 진행 상황 표시
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    # 임시 ZIP 파일 생성
                    zip_buffer = io.BytesIO()
                    
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        total_files = len(st.session_state.processed_files)
                        
                        for i, (filename, file_data) in enumerate(st.session_state.processed_files.items()):
                            status_text.text(f"처리 중: {filename}")
                            
                            # 각 파일 정렬 처리 (디버깅 출력 억제)
                            parsed_df = file_data['parsed_df'].copy()
                            sorted_df = apply_sorting_to_parsed_file_silent(parsed_df)
                            
                            # 메모리 내에서 엑셀 파일 생성 (색상 표시 포함)
                            excel_buffer = io.BytesIO()
                            
                            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                                sorted_df.to_excel(writer, sheet_name='파싱결과_정렬됨', index=False)
                                
                                # 색상 표시 적용
                                workbook = writer.book
                                worksheet = writer.sheets['파싱결과_정렬됨']
                                
                                combined_rows = find_combined_delivery_rows(sorted_df)
                                heavy_rows = find_heavy_order_rows(sorted_df)
                                
                                if combined_rows or heavy_rows:
                                    from openpyxl.styles import PatternFill
                                    
                                    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                                    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                                    red_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
                                    
                                    overlap_rows = list(set(combined_rows) & set(heavy_rows))
                                    
                                    for row_idx in range(len(sorted_df)):
                                        excel_row = row_idx + 2
                                        
                                        if row_idx in overlap_rows:
                                            fill = red_fill
                                        elif row_idx in combined_rows:
                                            fill = yellow_fill
                                        elif row_idx in heavy_rows:
                                            fill = green_fill
                                        else:
                                            continue
                                        
                                        for col_idx in range(1, len(sorted_df.columns) + 1):
                                            cell = worksheet.cell(row=excel_row, column=col_idx)
                                            cell.fill = fill
                            
                            # ZIP에 파일 추가 (메모리에서 직접)
                            zip_filename = f"{filename.split('.')[0]}_parsed_sorted.xlsx"
                            zipf.writestr(zip_filename, excel_buffer.getvalue())
                            
                            # 진행률 업데이트
                            progress_bar.progress((i + 1) / total_files)
                    
                    # ZIP 파일 완성
                    zip_buffer.seek(0)
                    
                    # 다운로드 버튼
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    zip_filename = f"전체_파싱결과_{timestamp}.zip"
                    
                    status_text.text("✅ ZIP 파일 생성 완료!")
                    progress_bar.progress(100)
                    
                    st.download_button(
                        label="💾 ZIP 파일 다운로드",
                        data=zip_buffer.getvalue(),
                        file_name=zip_filename,
                        mime="application/zip"
                    )
                    
                    st.success(f"✅ {len(st.session_state.processed_files)}개 파일이 ZIP으로 패키징되었습니다!")
                
                except Exception as e:
                    st.error(f"ZIP 생성 실패: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())
        
        # 패킹리스트 섹션
        st.header("📦 4. 패킹리스트 생성")
        
        col1, col2 = st.columns([2, 1])
        
        with col2:
            st.markdown("### 📊 현재 집계 현황")
            stats = st.session_state.packing_generator.get_statistics()
            
            st.metric("처리된 파일", stats['files_processed'])
            st.metric("전체 주문 건수", stats['total_orders_processed'])
            st.metric("상품 종류", stats['total_products'])
            st.metric("총 수량", stats['total_orders'])
            st.metric("총 무게", f"{stats['total_weight']:.1f} KG")
            
            # 배송 정보
            if stats['unique_deliveries'] > 0:
                st.markdown("#### 🚚 배송 정보")
                st.metric("고유 배송지", stats['unique_deliveries'])
                st.metric("합배송 건", stats['combined_deliveries'])
        
        with col1:
            if st.button("📦 패킹리스트 생성", type="primary"):
                try:
                    packing_items, summary = st.session_state.packing_generator.generate_packing_list()
                    
                    if not packing_items:
                        st.warning("패킹할 상품이 없습니다.")
                    else:
                        # 패킹리스트 결과 표시
                        st.success(f"✅ 패킹리스트 생성 완료! (총 {len(packing_items)}개 상품)")
                        
                        # 요약 정보
                        st.markdown("### 📋 패킹 요약")
                        summary_cols = st.columns(5)
                        
                        with summary_cols[0]:
                            st.metric("총 상품 종류", summary.total_items)
                        with summary_cols[1]:
                            st.metric("총 무게", f"{summary.total_weight:.1f} KG")
                        with summary_cols[2]:
                            st.metric("전체 주문 건수", summary.total_orders)
                        with summary_cols[3]:
                            st.metric("고유 배송지", summary.unique_delivery_locations)
                        with summary_cols[4]:
                            st.metric("합배송 건", summary.combined_delivery_count)
                        
                        # 패킹리스트 테이블
                        st.markdown("### 📦 상세 패킹리스트")
                        
                        # 간소화된 데이터프레임 생성
                        packing_data = []
                        for item in packing_items:
                            row = {
                                '상품명': item.product_name,
                            }
                            
                            # 수량 표시 (모든 수량 정수로 표시)
                            row['수량'] = int(item.quantity)
                            
                            # 총무게 (닭발 제외)
                            if item.category == '닭발':
                                row['총무게(KG)'] = ''  # 닭발은 무게 표시 안 함
                            elif item.total_weight > 0:
                                row['총무게(KG)'] = f"{item.total_weight:.1f}"
                            else:
                                row['총무게(KG)'] = ''
                            
                            packing_data.append(row)
                        
                        packing_df = pd.DataFrame(packing_data)
                        st.dataframe(packing_df, use_container_width=True)
                        
                        # ❌ 배송 분석 섹션 완전 제거 (UX 개선)
                        # 기존의 배송 분석 표시 코드를 모두 제거
                        
                        # 다운로드 버튼
                        st.markdown("### 💾 다운로드")
                        
                        col1, col2 = st.columns(2)
                        
                        # 엑셀 다운로드 - 패킹리스트만 (배송분석 시트 제거)
                        with col1:
                            output = io.BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                # 패킹리스트 시트만 생성 (배송분석 시트 제거)
                                packing_df.to_excel(writer, sheet_name='패킹리스트', index=False)
                            
                            output.seek(0)
                            
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            filename = f"패킹리스트_{timestamp}.xlsx"
                            
                            st.download_button(
                                label="📊 패킹리스트 엑셀 다운로드",
                                data=output.getvalue(),
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        
                        # CSV 다운로드
                        with col2:
                            csv_data = packing_df.to_csv(index=False, encoding='utf-8-sig')
                            
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            filename = f"패킹리스트_{timestamp}.csv"
                            
                            st.download_button(
                                label="📄 CSV 다운로드",
                                data=csv_data,
                                file_name=filename,
                                mime="text/csv"
                            )
                        
                except Exception as e:
                    st.error(f"패킹리스트 생성 실패: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())
    
    else:
        st.info("👆 먼저 엑셀 파일을 업로드하세요.")

if __name__ == "__main__":
    main()
